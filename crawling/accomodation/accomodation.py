import openpyxl
import requests
import pandas as pd
from itertools import islice
import atexit

fileName = 'Accomodations'
path = 'C:/Users/Chan/Desktop/' + fileName + '.xlsx'

places = {'Seoul' : '서울', 'Busan_Province' : '부산', 'Daegu_Metropolitan_City' : '대구', 'Incheon_Metropolitan_City' : '인천', 'Gwangju_Metropolitan_City' : '광주', 
        'Daejeon_Metropolitan_City' : '대전', 'Ulsan_Metropolitan_City' : '울산', 'Sejong' : '세종', 'Gangwon_Province' : '강원','Gyeonggi_Province' : '경기', 
        'South_Gyeongsang_Province' : '경남', 'North_Gyeongsang_Province' : '경북', 'South_Jeolla_Province' : '전남', 'North_Jeolla_Province' : '전북', 
        'South_Chungcheong_Province' : '충남', 'North_Chungcheong_Province' : '충북', 'Jeju' : '제주'}
catagory = {'propertyType' : '숙박유형', 'starRating' : '성급', 'name' : '이름', 'road' : '도로명주소', 'lot' : '지번주소', 'phone' : '전화번호', 'id' : '식별번호'}

try:
    wb = openpyxl.load_workbook(path, data_only = True)
except:
    wb = openpyxl.Workbook()
    del wb['Sheet']

    columnSize = [15, 5, 50, 60, 60, 15]

    for sheetName in places.values():
        ws = wb.create_sheet()
        ws.title = sheetName

        for i, c in enumerate(dict(islice(catagory.items(), len(catagory) - 1)).values()):
            ws.cell(row = 1, column = i + 1).value = c
            ws.column_dimensions[chr(i + 65)].width = columnSize[i]

    ws = wb.create_sheet()
    ws.title = catagory['id']

    wb.save(path)

    wb = openpyxl.load_workbook(path, data_only = True)

# insertion sort (오름차순)
def IS(id, col):
    for i, cp in enumerate(col):
        if cp >= id:
            col.insert(i, id)
            wsi.insert_rows(i + 1)
            wsi.cell(row = i + 1, column = 1).value = id
            return
    col.append(id)
    if len(col) - 1 == 0:
        wsi.cell(row = 1, column = 1).value = id
    else:
        wsi.cell(row = wsi.max_row + 1, column = 1).value = id

# 이진탐색
def BS(id, col):    # 있으면 True 없으면 False
    size = len(col)
    if not size:
        return 0

    st = 0
    ed = size - 1
    while st < ed:
        md = int((ed + st) / 2)
        if col[md] > id:
            ed = md - 1
        elif col[md] < id:
            st = md + 1
        else:
            return True

    if col[st] == id:
        return True
    else:
        return False

def sv():
    if endCheck == False:
        IS(id, ids)

    wb.save(path)
    wb.close()

class entity:
    def __init__(self):
        self.property = []

df = pd.read_excel(path, catagory['id'], header = None, index_col = None)
df = df.transpose()
idReq = df.values.tolist()
if len(idReq) >= 1:
    ids = idReq[0][ : ]
else:
    ids = []

wsi = wb[catagory['id']]

endCheck = True
atexit.register(sv)

headers = {'Content-Type': 'application/json; charset = utf-8'}

try:
    for place, placeK in places.items():
        print(placeK)

        ws = wb[placeK]
        lastRow = ws.max_row
        page = 0
        while 1:
            listUrl = 'https://m-hotel.naver.com/hotels/api/hotels?chains=&checkin=&checkout=&destination=place:' + place + '&features=&guestRatings=&includeLocalTaxesInTotal=false&includeTaxesInTotal=false&maxPrice=&minPrice=&pageIndex=' + str(page) + '&pageSize=100&propertyTypes=&radius=&rooms=&sortDirection=descending&sortField=popularityKR&starRating=&type=pc'
            listReq = requests.get(listUrl, headers = headers)
            listBody = listReq.json()

            results = listBody['results']
            resultsSize = len(results)
            if resultsSize == 0:
                break

            for rs in results:  # id Check
                id = rs['id']
                if BS(id, ids) == True:
                    continue

                we = entity()

                key = rs['key']
                resultUrl = 'https://m-hotel.naver.com/hotels/api/hotels/' + key + '?groupedFeatures=true&type=pc'
                resultReq = requests.get(resultUrl, headers = headers)
                resultBody = resultReq.json()

                for col in listBody['propertyTypes']:   # 숙박유형
                    pid = col['id']
                    name = col['name']
                    if pid == rs['propertyType']:
                        we.property.append(name)
                        if name == '호텔':
                            starR = rs['starRating']
                            if starR == 0:
                                starR = '-'
                            else:
                                starR = str(starR) + '성'
                        else:
                            starR = '-'

                        we.property.append(starR)
                        break

                we.property.append(rs['name']) # 이름

                pinID = None
                if resultBody.get('domestic') != None:  # pin ID
                    pinID = resultBody['domestic'].get('pinID')
                    if pinID != None:
                        mapUrl = 'https://map.naver.com/v5/api/sites/summary/' + pinID + '?lang=ko'
                        mapReq = requests.get(mapUrl, headers = headers)
                        mapBody = mapReq.json()

                        try:
                            roadAd = mapBody['roadAddr']['text']
                        except:
                            if placeK in dict(islice(places.items(), 7)).values():
                                roadAd = placeK + ' ' + rs['place']['name']
                            elif placeK == '세종':
                                roadAd = placeK
                            else:
                                roadAd = placeK + ' ' + rs['place']['name'] + '시(군)'

                            if rs.get('address') != None:
                                roadAd += ' ' + rs['address']

                        try:
                            lotAd = mapBody['address']
                        except:
                            lotAd = '-'
                        try:
                            phone = mapBody['phone']
                        except:
                            phone = '-'

                        we.property.append(roadAd)  # 도로명
                        we.property.append(lotAd)   # 지번
                        we.property.append(phone)   # 전화번호

                if pinID == None:
                    if placeK in dict(islice(places.items(), 7)).values():
                        roadAd = placeK + ' ' + rs['place']['name']
                    elif placeK == '세종':
                        roadAd = placeK
                    else:
                        roadAd = placeK + ' ' + rs['place']['name'] + '시(군)'

                    if rs.get('address') != None:
                        roadAd += ' ' + rs['address']

                    we.property.append(roadAd)
                    for _ in range(2):
                        we.property.append('-')

                for i, p in enumerate(we.property):
                    ws.cell(row = lastRow + 1, column = i + 1).value = p
                lastRow += 1

                wb.save(path)
                endCheck = False

                IS(id, ids)
                wb.save(path)

                endCheck = True
            
            page += 1

except:
    exit()
