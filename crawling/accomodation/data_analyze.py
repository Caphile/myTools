import openpyxl
import pandas as pd

path = 'Accomodations_final.xlsx'
wb = openpyxl.load_workbook(path, data_only = False)

provs = wb.get_sheet_names()[ : -1] # 식별번호 시트 사용하지 않음
stats = {'전체' : {}}
for prov in provs:
    stats[prov] = {}

for sheet_name in provs:
    ws = wb.get_sheet_by_name(sheet_name)

    for row in range (2, ws.max_row):
        accom_type = ws.cell(row, 1).value

        # 시.도별 통계
        if accom_type not in stats[sheet_name]:
            stats[sheet_name][accom_type] = 0
        stats[sheet_name][accom_type] += 1

        # 전체 통계
        if accom_type not in stats['전체']:
            stats['전체'][accom_type] = 0
        stats['전체'][accom_type] += 1

accom_types = list(stats['전체'].keys())

# 통계 저장용 데이터프레임 제작
data = [[0 for _ in range(len(provs) + 2)] for _ in range(len(accom_types) + 2)]
data[0][0] = '숙박/지역'
for idx, accom_type in enumerate(accom_types):
    data[idx + 1][0] = accom_type
for idx, prov in enumerate(provs):
    data[0][idx + 1] = prov
data[len(accom_types) + 1][0] = data[0][len(provs) + 1] = '합계'

# 데이터프레임에 데이터  채우기
for c in range(1, len(provs) + 1):
    prov = data[0][c]
    stat_by_prov = stats[prov]

    for r in range(1, len(accom_types) + 1):
        accom_type = data[r][0]
        if accom_type in stat_by_prov:
            data[r][c] = stat_by_prov[accom_type]

# 합계 저장
for c in range(1, len(provs) + 1):
    sum = 0
    for r in range(1, len(accom_types) + 1):
        sum += data[r][c]
    data[len(accom_types) + 1][c] = sum

for r in range(1, len(accom_types) + 2):
    sum = 0
    for c in range(1, len(provs) + 1):
        sum += data[r][c]
    data[r][len(provs) + 1] = sum

# 엑셀 저장
df = pd.DataFrame(data)
df.to_excel('stats.xlsx', index = False, header = False)