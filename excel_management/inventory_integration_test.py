from tkinter import filedialog
from tkinter import *
from tkinter.ttk import *
import openpyxl
import datetime

def btn_click(opt, user):
    st = fileRD(opt, user)
    if st:
        btn2['state'] = NORMAL

def fileRD(opt, user):
    global copydata
    welldone = True

    try:
        local = filedialog.askopenfilename(filetypes=(('Excel files','*.xlsx'),('all files','*.*')))
        wb = openpyxl.load_workbook(local)
        ws = wb.active

        user += 'test'
        if opt == 0:
            copydata = []
        else:
            count = 0
        for i in range(1, ws.max_column + 1):
            if ws.cell(row = 1, column = i).value == user:
                for j in range(2, ws.max_row + 1):
                    if opt == 0:
                        copydata.append(ws.cell(row = j, column = i).value)
                    else:
                        if copydata == []:
                            welldone = False
                            break
                        ws.cell(row = j, column = i, value = copydata[count])
                        count += 1
        if opt == 1:
            fileWR(wb)
        wb.close()
    except:
        welldone = False

    return welldone

def fileWR(workbook):
    dt = datetime.datetime.now()

    savename = '//192.168.0.180/Projects/구매자재/02.설계/DAOL_Pro_Inventory_통합테스트시나리오_'
    savename += str(dt.year) + str(dt.month).zfill(2) + str(dt.day).zfill(2) + '.xlsx'

    workbook.save(savename)

    win.destroy()
    win.quit()
    exit()

win = Tk()
win.geometry('300x200')
win.resizable(False, False)
win.title('엑셀통합관리')

selected_user = StringVar()
users = ['사업관리팀', 'QA']

rbs = []
for i in users:
    rb = Radiobutton(win, text = i, value = i, variable = selected_user)
    rbs.append(rb)
    rb.pack(fill = 'x', padx = 5, pady = 5, ipady = 5)

btn1 = Button(win, text = '복사', command = lambda : btn_click(0, selected_user.get()))
btn1.pack(fill = 'x', padx = 5, pady = 5, ipady = 5)

btn2 = Button(win, text = '붙여넣기', state = 'disabled', command = lambda : fileRD(1, selected_user.get()))
btn2.pack(fill = 'x', padx = 5, pady = 5, ipady = 5)

win.mainloop()
